#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2020/5/20 9:16
# @Author : Lemon_Tricy
# @QQ: 2378807189
# Copyright：湖南省零檬信息技术有限公司

'''
接口测试不步骤：
1、 接口文档--接口测试用例
2、 依据测试用例 执行测试 -- 数据
3、执行结束之后-- 执行结果 vs 预期结果 ==  通过  不通过  --写入测试用例

实现去Excel表格里读取数据分功能 --- 借助第三方库 == openpyxl库 ：读取数据 + 写入数据
1）  安装： pip install openpyxl
2） 导入Python 文件:

Excel表格三大对象：
1、工作簿对象
2、表单  --- 工作簿去找表单
3、 单元格  ---  行号  列号
cell = sheet.cell(row=1,column=1)  # 单元格
print(cell.value)  # 可以获取单元格的内容
cell.value = '测试编号'  # 重新赋值
# 只要写入Excel内容 --要保存  === 注意： 关闭文件
wb.save("test_case.xlsx")   # 保存文件
print(cell.value)

'''
import openpyxl  # 导入成功
import requests

# 读取测试用例数据的函数
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)  # 加载了这个工作簿  -- Excel表格 ==赋值给一个变量
    sheet = wb[sheetname]   # 表单
    max_row = sheet.max_row  # 获取最大行号
    cases = []  # 空列表
    for i in range(2,max_row+1):
        case = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,   # 获取url
        data = sheet.cell(row=i,column=6).value,  # 获取data
        expected_result = sheet.cell(row=i,column=7).value  # 获取期望结果
        )   # 一个用例存放到一个字典
        cases.append(case)   # 把字典追加到列表里保存起来
    return cases    # 定义成返回值
# read_data("test_case.xlsx","login")

# 写入结果的方法--函数
def write_result(filename,sheetname,row,column,real_result):
    wb = openpyxl.load_workbook(filename)  # 加载了这个工作簿  -- Excel表格 ==赋值给一个变量
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = real_result   # 写入
    wb.save(filename)

# 函数-- 发送接口请求
def post_func(qcd_url,qcd_data):
    res = requests.post(url=qcd_url,data=qcd_data)  # post方法发送接口请求
    result = res.json()   # 字典
    return result   # 返回值

